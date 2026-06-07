"""Gera a planilha EVOLUT x BYD - modelagem de negocio (xlsx com formulas vivas)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

# ---------- estilos ----------
BYD_RED = "EC1C24"
DARK = "1A1A1A"
YELLOW = "FFF2CC"
GREEN = "C6EFCE"
GREY = "F2F2F2"

title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
title_fill = PatternFill("solid", fgColor=DARK)
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor=BYD_RED)
input_fill = PatternFill("solid", fgColor=YELLOW)
total_font = Font(bold=True)
total_fill = PatternFill("solid", fgColor=GREY)
result_fill = PatternFill("solid", fgColor=GREEN)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

BRL = 'R$ #,##0.00'
PCT = '0.00%'
INT = '#,##0'

wb = Workbook()


def style_title(ws, cell_range, text):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.font = title_font
    c.fill = title_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[c.row].height = 24


def style_header_row(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        c = ws.cell(row=row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border


def footer_note(ws, row, col_span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1)
    c.value = ("Projecao baseada em dados publicos de mercado (Fenabrave/BYD 2025). "
               "Adesao e distribuicao por estado sao estimativas a validar em piloto.")
    c.font = Font(italic=True, size=9, color="808080")


# =====================================================================
# ABA 1 - PREMISSAS
# =====================================================================
ws1 = wb.active
ws1.title = "Premissas"
ws1.sheet_view.showGridLines = False

style_title(ws1, "A1:D1", "EVOLUT x BYD — PREMISSAS (PAINEL DE CONTROLE)")
ws1["A2"] = "Altere os valores em amarelo: toda a planilha recalcula automaticamente."
ws1["A2"].font = Font(italic=True, size=10)

style_header_row(ws1, 3, 1, 4)
ws1["A3"] = "Premissa"
ws1["B3"] = "Valor"
ws1["C3"] = "Unidade"
ws1["D3"] = "Observacao"

premissas = [
    # (label, value, unit, note, number_format)
    ("Total de concessionarias BYD no Brasil", 220, "unid.", "meta: 250 concessionarias", INT),
    ("Meta de concessionarias BYD", 250, "unid.", "expansao planejada", INT),
    ("BEVs vendidos/ano no Brasil (BYD, 2025)", 46000, "unid./ano", "lider com 73,62% do mercado BEV", INT),
    ("% das vendas BEV concentradas em SP", 0.25, "%", "maior concentracao: SP, DF, RJ, RS", PCT),
    ("Valor mensal SaaS por concessionaria", 800, "R$/mes", "linha de receita 1", BRL),
    ("Valor mensal por totem publico (itinerante)", 1500, "R$/mes", "shoppings, aeroportos, eventos", BRL),
    ("Valor por lead qualificado", 25, "R$/lead", "linha de receita 2", BRL),
    ("Valor da instalacao solar por kit (com imposto)", 2200, "R$/kit", "kit: 4 paineis + 1 microinversor (~2,2 kWp)", BRL),
    ("Carga tributaria embutida na instalacao", 0.15, "%", "imposto incluso no valor do kit", PCT),
    ("Custo repassado ao instalador credenciado por kit", 1500, "R$/kit", "repasse ao instalador acreditado", BRL),
    ("% adesao a instalacao solar — cenario conservador", 0.50, "%", "cenario 1 de 3", PCT),
    ("% adesao a instalacao solar — cenario realista", 0.65, "%", "cenario 2 de 3", PCT),
    ("% adesao a instalacao solar — cenario agressivo", 0.75, "%", "cenario 3 de 3", PCT),
    ("Taxa de plataforma de gestao mensal (BYD central)", 20000, "R$/mes", "linha de receita 5", BRL),
    ("Leads gerados por totem por mes", 150, "leads/totem/mes", "estimativa para calculo da linha de receita 2", INT),
]

start_row = 4
for i, (label, value, unit, note, fmt) in enumerate(premissas):
    r = start_row + i
    ws1.cell(row=r, column=1, value=label)
    cell = ws1.cell(row=r, column=2, value=value)
    cell.fill = input_fill
    cell.number_format = fmt
    cell.border = border
    ws1.cell(row=r, column=3, value=unit)
    ws1.cell(row=r, column=4, value=note)
    for col in (1, 3, 4):
        ws1.cell(row=r, column=col).border = border

# referencias nomeadas (linhas)
P = {
    "total_conc": "Premissas!$B$4",
    "meta_conc": "Premissas!$B$5",
    "bevs_brasil": "Premissas!$B$6",
    "pct_sp": "Premissas!$B$7",
    "saas_concessionaria": "Premissas!$B$8",
    "saas_totem": "Premissas!$B$9",
    "valor_lead": "Premissas!$B$10",
    "valor_kit": "Premissas!$B$11",
    "carga_tributaria": "Premissas!$B$12",
    "custo_instalador": "Premissas!$B$13",
    "adesao_conservador": "Premissas!$B$14",
    "adesao_realista": "Premissas!$B$15",
    "adesao_agressivo": "Premissas!$B$16",
    "taxa_plataforma": "Premissas!$B$17",
    "leads_por_totem": "Premissas!$B$18",
}

# Custos (DRE) - tambem como premissas editaveis
ws1.cell(row=20, column=1, value="Premissas de custos (DRE)").font = Font(bold=True, size=12, color=BYD_RED)
custos = [
    ("Custo de infraestrutura/hospedagem do SaaS (% da receita SaaS)", 0.05, "%", "servidores, cloud, manutencao de software", PCT),
    ("Custo de manutencao por totem (R$/totem/mes)", 150, "R$/totem/mes", "manutencao fisica e suporte tecnico", BRL),
    ("Custo de equipe de coordenacao (mensal)", 30000, "R$/mes", "equipe interna de operacao e parcerias", BRL),
    ("Impostos sobre receita de servicos (% da receita bruta)", 0.08, "%", "ISS/PIS/COFINS sobre servicos prestados", PCT),
    ("Marketing e aquisicao (% da receita bruta)", 0.04, "%", "campanhas, eventos, materiais", PCT),
]
start_row2 = 21
for i, (label, value, unit, note, fmt) in enumerate(custos):
    r = start_row2 + i
    ws1.cell(row=r, column=1, value=label)
    cell = ws1.cell(row=r, column=2, value=value)
    cell.fill = input_fill
    cell.number_format = fmt
    cell.border = border
    ws1.cell(row=r, column=3, value=unit)
    ws1.cell(row=r, column=4, value=note)
    for col in (1, 3, 4):
        ws1.cell(row=r, column=col).border = border

C = {
    "custo_infra_pct": "Premissas!$B$21",
    "custo_manut_totem": "Premissas!$B$22",
    "custo_equipe": "Premissas!$B$23",
    "imposto_servico_pct": "Premissas!$B$24",
    "marketing_pct": "Premissas!$B$25",
}

ws1.column_dimensions["A"].width = 56
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 48
ws1.freeze_panes = "A4"
footer_note(ws1, 28, 4)

# =====================================================================
# ABA 2 - CONCESSIONARIAS E TOTENS POR ESTADO
# =====================================================================
ws2 = wb.create_sheet("Concessionarias e Totens")
ws2.sheet_view.showGridLines = False
style_title(ws2, "A1:H1", "DISTRIBUICAO DE CONCESSIONARIAS E TOTENS POR ESTADO")

headers = ["Estado", "% estimado de mercado BYD", "Nº de concessionarias (estimado)",
           "Totens em concessionaria", "Totens itinerantes (estimado)",
           "Total de totens no estado", "Receita SaaS concessionarias/mes (R$)",
           "Receita totens itinerantes/mes (R$)"]
hr = 3
for i, h in enumerate(headers, start=1):
    ws2.cell(row=hr, column=i, value=h)
style_header_row(ws2, hr, 1, len(headers))

# distribuicao proporcional (soma = 100%) - SP maior fatia, demais conforme porte de mercado
estados = [
    ("Sao Paulo", 0.20),
    ("Rio de Janeiro", 0.09),
    ("Minas Gerais", 0.08),
    ("Parana", 0.08),
    ("Rio Grande do Sul", 0.08),
    ("Distrito Federal", 0.06),
    ("Bahia", 0.06),
    ("Santa Catarina", 0.05),
    ("Goias", 0.04),
    ("Pernambuco", 0.04),
    ("Ceara", 0.02),
    ("Espirito Santo", 0.02),
    ("Para", 0.01),
    ("Amazonas", 0.02),
    ("Mato Grosso", 0.02),
    ("Mato Grosso do Sul", 0.02),
    ("Paraiba", 0.01),
    ("Rio Grande do Norte", 0.01),
    ("Alagoas", 0.01),
    ("Sergipe", 0.01),
    ("Maranhao", 0.01),
    ("Piaui", 0.01),
    ("Tocantins", 0.01),
    ("Rondonia", 0.01),
    ("Acre", 0.01),
    ("Roraima", 0.01),
    ("Amapa", 0.01),
]
assert abs(sum(p for _, p in estados) - 1.0) < 1e-9

first_data_row = hr + 1
last_data_row = first_data_row + len(estados) - 1
for i, (uf, pct) in enumerate(estados):
    r = first_data_row + i
    ws2.cell(row=r, column=1, value=uf).border = border
    c = ws2.cell(row=r, column=2, value=pct)
    c.number_format = PCT
    c.fill = input_fill
    c.border = border
    # nº concessionarias = arredondamento(total_conc * %)
    c3 = ws2.cell(row=r, column=3, value=f"=ROUND({P['total_conc']}*B{r},0)")
    c3.number_format = INT
    c3.border = border
    # totens em concessionaria = 1 por concessionaria
    c4 = ws2.cell(row=r, column=4, value=f"=C{r}")
    c4.number_format = INT
    c4.border = border
    # totens itinerantes = 1 a cada 3 concessionarias
    c5 = ws2.cell(row=r, column=5, value=f"=ROUND(C{r}/3,0)")
    c5.number_format = INT
    c5.border = border
    # total de totens
    c6 = ws2.cell(row=r, column=6, value=f"=D{r}+E{r}")
    c6.number_format = INT
    c6.border = border
    # receita saas concessionarias / mes
    c7 = ws2.cell(row=r, column=7, value=f"=C{r}*{P['saas_concessionaria']}")
    c7.number_format = BRL
    c7.border = border
    # receita totens itinerantes / mes
    c8 = ws2.cell(row=r, column=8, value=f"=E{r}*{P['saas_totem']}")
    c8.number_format = BRL
    c8.border = border

total_row = last_data_row + 1
ws2.cell(row=total_row, column=1, value="TOTAL").font = total_font
for col, letter in [(2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G"), (8, "H")]:
    cell = ws2.cell(row=total_row, column=col,
                    value=f"=SUM({letter}{first_data_row}:{letter}{last_data_row})")
    cell.font = total_font
    cell.fill = total_fill
    cell.border = border
    if col == 2:
        cell.number_format = PCT
    elif col in (3, 4, 5, 6):
        cell.number_format = INT
    else:
        cell.number_format = BRL

# referencias para outras abas
TOT_CONC = f"'Concessionarias e Totens'!$C${total_row}"
TOT_TOTENS = f"'Concessionarias e Totens'!$F${total_row}"
TOT_TOTENS_ITIN = f"'Concessionarias e Totens'!$E${total_row}"

note_row = total_row + 2
ws2.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
ws2.cell(row=note_row, column=1,
         value=("Nota: totens itinerantes circulam por shoppings, aeroportos e eventos de grande fluxo, "
                "ampliando a presenca da marca alem das concessionarias fixas."))
ws2.cell(row=note_row, column=1).font = Font(italic=True, size=9)
ws2.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)

for col, width in [(1, 22), (2, 16), (3, 18), (4, 14), (5, 16), (6, 14), (7, 22), (8, 22)]:
    ws2.column_dimensions[get_column_letter(col)].width = width
ws2.freeze_panes = "A4"
footer_note(ws2, note_row + 2, 8)

# =====================================================================
# ABA 3 - INSTALACOES SOLARES (SP)
# =====================================================================
ws3 = wb.create_sheet("Instalacoes Solares (SP)")
ws3.sheet_view.showGridLines = False
style_title(ws3, "A1:D1", "INSTALACOES SOLARES — VOLUME E RENTABILIDADE (SAO PAULO)")

ws3["A3"] = "BEVs vendidos/ano no Brasil"
ws3["B3"] = f"={P['bevs_brasil']}"
ws3["B3"].number_format = INT
ws3["A4"] = "% das vendas BEV concentradas em SP"
ws3["B4"] = f"={P['pct_sp']}"
ws3["B4"].number_format = PCT
ws3["A5"] = "BEVs vendidos/ano em SP"
ws3["B5"] = "=B3*B4"
ws3["B5"].number_format = INT
ws3["A5"].font = Font(bold=True)
ws3["B5"].font = Font(bold=True)
for r in (3, 4, 5):
    ws3[f"A{r}"].border = border
    ws3[f"B{r}"].border = border

BEVS_SP = "$B$5"

scen_row = 7
ws3.cell(row=scen_row, column=1, value="Indicador (cenarios de adesao a instalacao solar)")
scenarios = [
    ("Conservador (50%)", P["adesao_conservador"], 2),
    ("Realista (65%)", P["adesao_realista"], 3),
    ("Agressivo (75%)", P["adesao_agressivo"], 4),
]
for name, _, col in scenarios:
    ws3.cell(row=scen_row, column=col, value=name)
style_header_row(ws3, scen_row, 1, 4)

rows_def = [
    ("% de adesao a instalacao solar", "pct_adesao", PCT),
    ("Instalacoes/ano", "instalacoes_ano", INT),
    ("Instalacoes/mes", "instalacoes_mes", INT),
    ("Receita bruta de instalacao (R$/ano)", "receita_bruta", BRL),
    ("Imposto embutido na instalacao (R$/ano)", "imposto", BRL),
    ("Custo repassado ao instalador (R$/ano)", "custo_instalador_total", BRL),
    ("Margem liquida por kit (R$)", "margem_kit", BRL),
    ("Receita liquida total no cenario (R$/ano)", "receita_liquida", BRL),
]

r0 = scen_row + 1
labels_rows = {}
for i, (label, key, fmt) in enumerate(rows_def):
    r = r0 + i
    ws3.cell(row=r, column=1, value=label).border = border
    labels_rows[key] = r
    for col in (2, 3, 4):
        ws3.cell(row=r, column=col).border = border
        ws3.cell(row=r, column=col).number_format = fmt

for name, adesao_ref, col in scenarios:
    L = get_column_letter(col)
    ws3[f"{L}{labels_rows['pct_adesao']}"] = f"={adesao_ref}"
    ws3[f"{L}{labels_rows['instalacoes_ano']}"] = f"=ROUND({BEVS_SP}*{L}{labels_rows['pct_adesao']},0)"
    ws3[f"{L}{labels_rows['instalacoes_mes']}"] = f"={L}{labels_rows['instalacoes_ano']}/12"
    ws3[f"{L}{labels_rows['receita_bruta']}"] = f"={L}{labels_rows['instalacoes_ano']}*{P['valor_kit']}"
    ws3[f"{L}{labels_rows['imposto']}"] = f"={L}{labels_rows['instalacoes_ano']}*{P['valor_kit']}*{P['carga_tributaria']}"
    ws3[f"{L}{labels_rows['custo_instalador_total']}"] = f"={L}{labels_rows['instalacoes_ano']}*{P['custo_instalador']}"
    ws3[f"{L}{labels_rows['margem_kit']}"] = f"={P['valor_kit']}-({P['valor_kit']}*{P['carga_tributaria']})-{P['custo_instalador']}"
    ws3[f"{L}{labels_rows['receita_liquida']}"] = f"={L}{labels_rows['instalacoes_ano']}*{L}{labels_rows['margem_kit']}"
    ws3[f"{L}{labels_rows['receita_liquida']}"].font = Font(bold=True)
    ws3[f"{L}{labels_rows['receita_liquida']}"].fill = result_fill

# referencias para outras abas (margem liquida da instalacao, por cenario)
SOLAR_LIQ = {
    "conservador": f"'Instalacoes Solares (SP)'!$B${labels_rows['receita_liquida']}",
    "realista": f"'Instalacoes Solares (SP)'!$C${labels_rows['receita_liquida']}",
    "agressivo": f"'Instalacoes Solares (SP)'!$D${labels_rows['receita_liquida']}",
}

ws3.column_dimensions["A"].width = 42
for col in ("B", "C", "D"):
    ws3.column_dimensions[col].width = 20
ws3.freeze_panes = "A8"
footer_note(ws3, labels_rows["receita_liquida"] + 3, 4)

# =====================================================================
# ABA 4 - RECEITA CONSOLIDADA (DRE simplificado)
# =====================================================================
ws4 = wb.create_sheet("Receita Consolidada (DRE)")
ws4.sheet_view.showGridLines = False
style_title(ws4, "A1:D1", "RECEITA CONSOLIDADA — DRE SIMPLIFICADO ANUAL (3 CENARIOS)")

hdr_row = 3
ws4.cell(row=hdr_row, column=1, value="Linha")
ws4.cell(row=hdr_row, column=2, value="Conservador (adesao 50%)")
ws4.cell(row=hdr_row, column=3, value="Realista (adesao 65%)")
ws4.cell(row=hdr_row, column=4, value="Agressivo (adesao 75%)")
style_header_row(ws4, hdr_row, 1, 4)

solar_cols = {2: SOLAR_LIQ["conservador"], 3: SOLAR_LIQ["realista"], 4: SOLAR_LIQ["agressivo"]}

rows4 = [
    "1. Receita SaaS concessionarias (anual)",
    "2. Receita SaaS totens itinerantes (anual)",
    "3. Receita por taxa de leads (anual)",
    "4. Margem liquida instalacoes solares (anual)",
    "5. Receita taxa de plataforma de gestao (anual)",
    "RECEITA BRUTA TOTAL",
    "",
    "Custo: infraestrutura/hospedagem SaaS",
    "Custo: manutencao de totens",
    "Custo: equipe de coordenacao",
    "Custo: impostos sobre receita de servicos",
    "Custo: marketing e aquisicao",
    "CUSTOS TOTAIS",
    "",
    "RESULTADO LIQUIDO ESTIMADO",
    "Margem liquida (%)",
]
r0 = hdr_row + 1
rr = {}
for i, label in enumerate(rows4):
    r = r0 + i
    ws4.cell(row=r, column=1, value=label)
    rr[label] = r

# linha 1: SaaS concessionarias anual = total_conc * valor_saas_conc * 12
ws4[f"B{rr['1. Receita SaaS concessionarias (anual)']}"] = f"={TOT_CONC}*{P['saas_concessionaria']}*12"
# linha 2: SaaS totens itinerantes anual
ws4[f"B{rr['2. Receita SaaS totens itinerantes (anual)']}"] = f"={TOT_TOTENS_ITIN}*{P['saas_totem']}*12"
# linha 3: leads anual = total_totens * leads_por_totem * 12 * valor_lead
ws4[f"B{rr['3. Receita por taxa de leads (anual)']}"] = f"={TOT_TOTENS}*{P['leads_por_totem']}*12*{P['valor_lead']}"
# linha 5: plataforma de gestao anual
ws4[f"B{rr['5. Receita taxa de plataforma de gestao (anual)']}"] = f"={P['taxa_plataforma']}*12"

# replicar linhas 1,2,3,5 nas 3 colunas (mesmo valor, formula referenciando coluna B)
for label in ["1. Receita SaaS concessionarias (anual)",
              "2. Receita SaaS totens itinerantes (anual)",
              "3. Receita por taxa de leads (anual)",
              "5. Receita taxa de plataforma de gestao (anual)"]:
    r = rr[label]
    ws4[f"C{r}"] = f"=$B{r}"
    ws4[f"D{r}"] = f"=$B{r}"

# linha 4: margem liquida instalacoes solares (difere por cenario)
r4 = rr["4. Margem liquida instalacoes solares (anual)"]
for col, ref in solar_cols.items():
    ws4.cell(row=r4, column=col, value=f"={ref}")

# RECEITA BRUTA TOTAL
r_bruta = rr["RECEITA BRUTA TOTAL"]
for col_letter in ("B", "C", "D"):
    ws4[f"{col_letter}{r_bruta}"] = (f"=SUM({col_letter}{rr['1. Receita SaaS concessionarias (anual)']}:"
                                     f"{col_letter}{rr['5. Receita taxa de plataforma de gestao (anual)']})")
    ws4[f"{col_letter}{r_bruta}"].font = total_font
    ws4[f"{col_letter}{r_bruta}"].fill = total_fill

# custos
r_infra = rr["Custo: infraestrutura/hospedagem SaaS"]
r_manut = rr["Custo: manutencao de totens"]
r_equipe = rr["Custo: equipe de coordenacao"]
r_imp_serv = rr["Custo: impostos sobre receita de servicos"]
r_mkt = rr["Custo: marketing e aquisicao"]
r_custos_tot = rr["CUSTOS TOTAIS"]
r_resultado = rr["RESULTADO LIQUIDO ESTIMADO"]
r_margem_pct = rr["Margem liquida (%)"]

r1 = rr["1. Receita SaaS concessionarias (anual)"]
r2 = rr["2. Receita SaaS totens itinerantes (anual)"]

for col_letter in ("B", "C", "D"):
    # infra = % * (receita saas concessionarias + receita saas totens)
    ws4[f"{col_letter}{r_infra}"] = f"={C['custo_infra_pct']}*({col_letter}{r1}+{col_letter}{r2})"
    # manutencao totens = total_totens * custo/totem/mes * 12
    ws4[f"{col_letter}{r_manut}"] = f"={TOT_TOTENS}*{C['custo_manut_totem']}*12"
    # equipe coordenacao
    ws4[f"{col_letter}{r_equipe}"] = f"={C['custo_equipe']}*12"
    # impostos sobre receita de servicos = % * receita bruta total
    ws4[f"{col_letter}{r_imp_serv}"] = f"={C['imposto_servico_pct']}*{col_letter}{r_bruta}"
    # marketing = % * receita bruta total
    ws4[f"{col_letter}{r_mkt}"] = f"={C['marketing_pct']}*{col_letter}{r_bruta}"
    # custos totais
    ws4[f"{col_letter}{r_custos_tot}"] = f"=SUM({col_letter}{r_infra}:{col_letter}{r_mkt})"
    ws4[f"{col_letter}{r_custos_tot}"].font = total_font
    ws4[f"{col_letter}{r_custos_tot}"].fill = total_fill
    # resultado liquido
    ws4[f"{col_letter}{r_resultado}"] = f"={col_letter}{r_bruta}-{col_letter}{r_custos_tot}"
    ws4[f"{col_letter}{r_resultado}"].font = Font(bold=True)
    ws4[f"{col_letter}{r_resultado}"].fill = result_fill
    # margem %
    ws4[f"{col_letter}{r_margem_pct}"] = f"={col_letter}{r_resultado}/{col_letter}{r_bruta}"

# formatos
for label, fmt in [(l, BRL) for l in rows4 if l and not l.startswith("Margem")]:
    r = rr[label]
    for col_letter in ("B", "C", "D"):
        cell = ws4[f"{col_letter}{r}"]
        if cell.value is not None:
            cell.number_format = BRL
            cell.border = border
ws4[f"B{r_margem_pct}"].number_format = PCT
ws4[f"C{r_margem_pct}"].number_format = PCT
ws4[f"D{r_margem_pct}"].number_format = PCT
for col_letter in ("B", "C", "D"):
    ws4[f"{col_letter}{r_margem_pct}"].border = border
    ws4[f"{col_letter}{r_margem_pct}"].font = Font(bold=True)

ws4.column_dimensions["A"].width = 46
for col in ("B", "C", "D"):
    ws4.column_dimensions[col].width = 24
ws4.freeze_panes = "A4"
footer_note(ws4, r_margem_pct + 2, 4)

# referencias para o dashboard (cenario realista = coluna C)
DRE_REF = {
    "saas_conc": f"'Receita Consolidada (DRE)'!$C${r1}",
    "saas_totem": f"'Receita Consolidada (DRE)'!$C${r2}",
    "leads": f"'Receita Consolidada (DRE)'!$C${rr['3. Receita por taxa de leads (anual)']}",
    "solar": f"'Receita Consolidada (DRE)'!$C${r4}",
    "plataforma": f"'Receita Consolidada (DRE)'!$C${rr['5. Receita taxa de plataforma de gestao (anual)']}",
    "bruta": f"'Receita Consolidada (DRE)'!$C${r_bruta}",
    "resultado": f"'Receita Consolidada (DRE)'!$C${r_resultado}",
    "resultado_cons": f"'Receita Consolidada (DRE)'!$B${r_resultado}",
    "resultado_real": f"'Receita Consolidada (DRE)'!$C${r_resultado}",
    "resultado_agr": f"'Receita Consolidada (DRE)'!$D${r_resultado}",
}

# =====================================================================
# ABA 5 - RESUMO EXECUTIVO (DASHBOARD)
# =====================================================================
ws5 = wb.create_sheet("Resumo Executivo")
ws5.sheet_view.showGridLines = False
style_title(ws5, "A1:F1", "RESUMO EXECUTIVO — EVOLUT x BYD (CENARIO REALISTA: ADESAO 65%)")

# cards
card_defs = [
    ("Receita bruta total anual", DRE_REF["bruta"], BRL),
    ("Nº de concessionarias", TOT_CONC, INT),
    ("Nº total de totens", TOT_TOTENS, INT),
    ("Instalacoes solares/ano", f"'Instalacoes Solares (SP)'!$C${labels_rows['instalacoes_ano']}", INT),
    ("Resultado liquido estimado (anual)", DRE_REF["resultado"], BRL),
]
card_row = 3
for i, (label, ref, fmt) in enumerate(card_defs):
    col = 1 + i
    c_label = ws5.cell(row=card_row, column=col, value=label)
    c_label.font = Font(bold=True, size=9, color="FFFFFF")
    c_label.fill = PatternFill("solid", fgColor=BYD_RED)
    c_label.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    c_val = ws5.cell(row=card_row + 1, column=col, value=f"={ref}")
    c_val.font = Font(bold=True, size=14)
    c_val.number_format = fmt
    c_val.alignment = Alignment(horizontal="center")
    c_val.fill = PatternFill("solid", fgColor=GREY)
    ws5.column_dimensions[get_column_letter(col)].width = 20
ws5.row_dimensions[card_row].height = 30
ws5.row_dimensions[card_row + 1].height = 26

# tabela: receita por linha (5 fontes) - cenario realista
tbl1_row = card_row + 4
ws5.cell(row=tbl1_row, column=1, value="Receita por linha (cenario realista, anual)").font = Font(bold=True, size=12, color=BYD_RED)
hdr2 = tbl1_row + 1
ws5.cell(row=hdr2, column=1, value="Linha de receita")
ws5.cell(row=hdr2, column=2, value="Valor anual (R$)")
style_header_row(ws5, hdr2, 1, 2)

linhas_receita = [
    ("1. SaaS concessionarias", DRE_REF["saas_conc"]),
    ("2. SaaS totens itinerantes", DRE_REF["saas_totem"]),
    ("3. Taxa por leads", DRE_REF["leads"]),
    ("4. Margem liquida instalacoes solares", DRE_REF["solar"]),
    ("5. Taxa de plataforma de gestao", DRE_REF["plataforma"]),
]
r0b = hdr2 + 1
for i, (label, ref) in enumerate(linhas_receita):
    r = r0b + i
    ws5.cell(row=r, column=1, value=label).border = border
    c = ws5.cell(row=r, column=2, value=f"={ref}")
    c.number_format = BRL
    c.border = border
last_lr_row = r0b + len(linhas_receita) - 1

# tabela: comparacao dos 3 cenarios
tbl2_row = last_lr_row + 3
ws5.cell(row=tbl2_row, column=1, value="Comparacao de cenarios — resultado liquido anual").font = Font(bold=True, size=12, color=BYD_RED)
hdr3 = tbl2_row + 1
ws5.cell(row=hdr3, column=1, value="Cenario")
ws5.cell(row=hdr3, column=2, value="Resultado liquido (R$)")
style_header_row(ws5, hdr3, 1, 2)
cenarios_dash = [
    ("Conservador (adesao 50%)", DRE_REF["resultado_cons"]),
    ("Realista (adesao 65%)", DRE_REF["resultado_real"]),
    ("Agressivo (adesao 75%)", DRE_REF["resultado_agr"]),
]
r0c = hdr3 + 1
for i, (label, ref) in enumerate(cenarios_dash):
    r = r0c + i
    ws5.cell(row=r, column=1, value=label).border = border
    c = ws5.cell(row=r, column=2, value=f"={ref}")
    c.number_format = BRL
    c.border = border
last_cen_row = r0c + len(cenarios_dash) - 1

# ----- graficos -----
# 1) barras: receita por linha
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Receita por linha (cenario realista)"
chart1.y_axis.title = "R$ / ano"
data1 = Reference(ws5, min_col=2, min_row=hdr2, max_row=last_lr_row)
cats1 = Reference(ws5, min_col=1, min_row=r0b, max_row=last_lr_row)
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.height = 8
chart1.width = 14
chart1.legend = None
ws5.add_chart(chart1, f"D{tbl1_row}")

# 2) barras: comparacao cenarios
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Comparacao de cenarios — resultado liquido anual"
chart2.y_axis.title = "R$ / ano"
data2 = Reference(ws5, min_col=2, min_row=hdr3, max_row=last_cen_row)
cats2 = Reference(ws5, min_col=1, min_row=r0c, max_row=last_cen_row)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.height = 8
chart2.width = 14
chart2.legend = None
ws5.add_chart(chart2, f"D{tbl2_row}")

# 3) pizza: composicao da receita
chart3 = PieChart()
chart3.title = "Composicao da receita (cenario realista)"
data3 = Reference(ws5, min_col=2, min_row=hdr2, max_row=last_lr_row)
cats3 = Reference(ws5, min_col=1, min_row=r0b, max_row=last_lr_row)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.height = 8
chart3.width = 14
ws5.add_chart(chart3, f"D{tbl1_row + 17}")

ws5.column_dimensions["A"].width = 34
ws5.column_dimensions["B"].width = 22
footer_note(ws5, last_cen_row + 19, 6)

# =====================================================================
out_path = "/home/user/evolut-byd/EVOLUT_x_BYD_Modelagem_de_Negocio.xlsx"
wb.save(out_path)
print("Arquivo salvo em:", out_path)
